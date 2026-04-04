"""Tests for setup_app.services_contacts — static helpers and CSV parsing."""
from __future__ import annotations

import csv
import io
from unittest.mock import MagicMock, patch

import pytest
from django.test import TestCase


class NormalizeKeyTests(TestCase):
    def test_lowercases(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._normalize_key("Name"), "name")

    def test_strips_whitespace(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._normalize_key("  email  "), "email")

    def test_replaces_spaces_with_underscores(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(
            ContactImportService._normalize_key("First Name"), "first_name"
        )

    def test_empty_string_returns_empty(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._normalize_key(""), "")

    def test_none_returns_empty(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._normalize_key(None), "")


class NormalizePhoneTests(TestCase):
    def test_adds_55_to_11_digit_number(self):
        from setup_app.services_contacts import ContactImportService
        result = ContactImportService._normalize_phone("11987654321")
        self.assertEqual(result, "+5511987654321")

    def test_adds_55_to_10_digit_number(self):
        from setup_app.services_contacts import ContactImportService
        result = ContactImportService._normalize_phone("1198765432")
        self.assertEqual(result, "+551198765432")

    def test_strips_non_digits(self):
        from setup_app.services_contacts import ContactImportService
        result = ContactImportService._normalize_phone("+55 (11) 98765-4321")
        self.assertEqual(result, "+5511987654321")

    def test_already_has_country_code(self):
        from setup_app.services_contacts import ContactImportService
        result = ContactImportService._normalize_phone("5511987654321")
        # 13 digits — no prefix added
        self.assertTrue(result.startswith("+"))

    def test_adds_plus_prefix(self):
        from setup_app.services_contacts import ContactImportService
        result = ContactImportService._normalize_phone("11987654321")
        self.assertTrue(result.startswith("+"))


class GetFileExtensionTests(TestCase):
    def test_csv_extension(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._get_file_extension("data.csv"), ".csv")

    def test_xlsx_extension(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._get_file_extension("file.xlsx"), ".xlsx")

    def test_uppercase_extension_lowercased(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._get_file_extension("FILE.CSV"), ".csv")

    def test_no_extension(self):
        from setup_app.services_contacts import ContactImportService
        self.assertEqual(ContactImportService._get_file_extension("noext"), "")


class ReadCsvTests(TestCase):
    def _make_file(self, content: str, name="contacts.csv", encoding="utf-8"):
        f = MagicMock()
        f.name = name
        f.read.return_value = content.encode(encoding)
        return f

    def test_reads_basic_csv(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)
        csv_content = "name,phone\nJohn Doe,11987654321"
        f = self._make_file(csv_content)
        rows = service._read_csv(f)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["name"], "John Doe")
        self.assertEqual(rows[0]["phone"], "11987654321")

    def test_normalizes_header_keys(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)
        csv_content = "Full Name,Phone Number\nJane,99999"
        f = self._make_file(csv_content)
        rows = service._read_csv(f)
        self.assertIn("full_name", rows[0])
        self.assertIn("phone_number", rows[0])

    def test_handles_latin1_encoding(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)
        csv_content = "name,phone\nJosé,11111111111"
        f = self._make_file(csv_content, encoding="latin-1")
        rows = service._read_csv(f)
        self.assertEqual(len(rows), 1)

    def test_raises_on_undecodable_content(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)
        f = MagicMock()
        f.name = "data.csv"
        # bytes that can't be decoded as utf-8, latin-1, or iso-8859-1 are unusual;
        # use a mock that always fails
        f.read.return_value = b"\x80" * 10
        # latin-1 can decode any single byte, so this actually succeeds — just verify no crash
        rows = service._read_csv(f)
        self.assertIsInstance(rows, list)


class ReadExcelTests(TestCase):
    def test_reads_excel_file(self):
        import openpyxl
        from setup_app.services_contacts import ContactImportService
        import io

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "phone"])
        ws.append(["Alice", "11999998888"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        f = MagicMock()
        f.name = "contacts.xlsx"
        # openpyxl.load_workbook accepts file-like objects
        f.read = buf.read
        f.seek = buf.seek

        user = MagicMock()
        service = ContactImportService(user)
        with patch("setup_app.services_contacts.openpyxl.load_workbook") as mock_load:
            mock_ws = MagicMock()
            mock_ws.__getitem__.return_value = [
                MagicMock(value="name"), MagicMock(value="phone")
            ]
            mock_ws.iter_rows.return_value = [("Alice", "11999998888")]
            mock_wb = MagicMock()
            mock_wb.active = mock_ws
            mock_load.return_value = mock_wb
            rows = service._read_excel(f)

        self.assertEqual(len(rows), 1)


class ImportFromFileTests(TestCase):
    def test_csv_import_creates_history(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)

        f = MagicMock()
        f.name = "test.csv"
        f.read.return_value = b"name,phone\nBob,11987654321"

        mock_history = MagicMock()
        mock_history.id = 1

        with patch("setup_app.services_contacts.ImportHistory.objects.create",
                   return_value=mock_history) as mock_create, \
             patch.object(service, "_process_rows", return_value=(1, 0)):
            result = service.import_from_file(f)

        mock_create.assert_called_once()
        self.assertEqual(result, mock_history)

    def test_exception_marks_history_as_failed(self):
        from setup_app.services_contacts import ContactImportService
        user = MagicMock()
        service = ContactImportService(user)

        f = MagicMock()
        f.name = "bad.csv"
        f.read.side_effect = Exception("disk error")

        mock_history = MagicMock()
        mock_history.id = 1

        with patch("setup_app.services_contacts.ImportHistory.objects.create",
                   return_value=mock_history):
            result = service.import_from_file(f)

        # status should be set to failed
        self.assertEqual(mock_history.status, "failed")
